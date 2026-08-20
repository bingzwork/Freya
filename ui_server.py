import argparse
import json
import mimetypes
import os
import queue
import re
import shutil
import subprocess
import threading
import time
import traceback
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FutureTimeoutError
import uuid
from pathlib import Path
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from urllib.parse import parse_qs, urlparse
from main import FreyaApp
from app.core.initializer import SystemConfig
from app.capabilities.formatter import format_capability_result
from app.capabilities.router import CapabilityResult
from app.core.request_context import RequestContext
from app.core.priority_llm import LLMPriority
from app.ui.agent_console import get_agent_console_snapshot, get_autonomy_snapshot, get_tasks_snapshot, get_memory_snapshot, get_system_snapshot
from app.research.intelligence import RequestSemanticAnalyzer, ResearchIntent
from app.research.task_learning import ResearchTaskSemanticAnalyzer, ResearchTaskLearningOrchestrator
FREYA=None
SUBSCRIBERS=set()
LAST_IMAGE_SUBJECT = ""
LAST_IMAGE_RESULT_URLS = []
UI_SESSION_ID = f"session_{uuid.uuid4().hex}"
UI_REQUEST_TIMEOUT_SECONDS = max(30.0, float(os.getenv("FREYA_UI_REQUEST_TIMEOUT_SECONDS", "180")))
RESEARCH_REQUEST_TIMEOUT_SECONDS = max(45.0, float(os.getenv("FREYA_RESEARCH_REQUEST_TIMEOUT_SECONDS", "180")))
DIRECT_CHAT_TIMEOUT_SECONDS = max(30.0, float(os.getenv("FREYA_DIRECT_CHAT_TIMEOUT_SECONDS", "90")))
BROWSER_ACTION_TIMEOUT_SECONDS = max(15.0, float(os.getenv("FREYA_BROWSER_ACTION_TIMEOUT_SECONDS", "45")))
LOCK=threading.Lock()
SHOPPING_STATE_LOCK=threading.RLock()
SHOPPING_STATES={}
WEB_SEARCH_SETTINGS_LOCK=threading.RLock()
WEB_SEARCH_SETTINGS={
    "enabled": os.getenv("FREYA_WEB_SEARCH_ENABLED", "true").lower() not in {"0", "false", "off", "no"},
    "provider": os.getenv("FREYA_WEB_SEARCH_PROVIDER", "exa").strip().lower() or "exa",
    "searxng_url": os.getenv("FREYA_SEARXNG_URL", "").strip(),
}
SUPPORTED_EXTENSIONS={".jpg",".jpeg",".png",".webp",".mp3",".wav",".m4a",".flac",".mp4",".mov",".webm",".txt",".md",".pdf",".docx",".csv",".xlsx",".json"}
IMAGE_EXTENSIONS={".jpg",".jpeg",".png",".webp",".gif",".bmp"}
RESEARCH_WORDS=("research","search","look this up","find information","deep web","latest","recent","current")
UNKNOWN_PERSON_REQUEST=re.compile(r"\b(find|search|latest|recent|current|look\s+up)\b.{0,80}\b(photo|picture|image)\b.{0,80}\b(this person|the person|him|her|them)\b",re.I)

def _web_search_settings_payload():
    with WEB_SEARCH_SETTINGS_LOCK:
        values = dict(WEB_SEARCH_SETTINGS)
    return {
        "enabled": bool(values.get("enabled", True)),
        "provider": str(values.get("provider") or "exa"),
        "searxng_url": str(values.get("searxng_url") or ""),
        "exa_api_key_configured": bool(os.getenv("EXA_API_KEY", "").strip()),
    }


def emit_avatar(state,**metadata):
    payload={"state":state,**metadata}
    with LOCK: subscribers=list(SUBSCRIBERS)
    for subscriber in subscribers:
        try: subscriber.put_nowait(payload)
        except queue.Full: pass

_INTERNAL_CHAT_LEAK_RE = re.compile(r"(?:success\s*:\s*(?:true|false)|answerability|context\s+evaluated|retrieved\s+\d+\s+results|no\s+provider\s+is\s+configured|found\s+\d+\s+tasks?|lesson\s+\[|candidate_type|workflow\s+validation|traceback|httperror)", re.I)

def _looks_like_internal_chat_leak(answer):
    text = str(answer or "").strip()
    return bool(text and _INTERNAL_CHAT_LEAK_RE.search(text))

def _safe_direct_local_chat(message, context=None):
    system = getattr(FREYA, "system", None) if FREYA is not None else None
    priority = getattr(system, "priority_llm", None)
    if priority is None:
        return "I couldn't complete that request reliably because Freya's local chat model is unavailable."
    try:
        outcome = priority.ask_outcome(
            prompt=str(message or "").strip(),
            system=("You are Freya, a helpful local AI assistant. Answer the user's request directly, naturally, and usefully. "
                    "Do not expose routing, memory, provider, capability, workflow, or internal diagnostic details."),
            priority=LLMPriority.CHAT,
            timeout=DIRECT_CHAT_TIMEOUT_SECONDS,
        )
        if getattr(outcome, "is_success", False) and str(getattr(outcome, "content", "") or "").strip():
            content = str(outcome.content).strip()
            if not _looks_like_internal_chat_leak(content):
                return content
    except Exception:
        pass
    return "I couldn't complete that request reliably because Freya's local chat model did not return a usable answer."

def _sanitize_chat_answer(answer, message, context=None):
    text = str(answer or "").strip()
    if _looks_like_internal_chat_leak(text):
        return _safe_direct_local_chat(message, context)
    return text

def _ffprobe(path):
    executable=shutil.which("ffprobe") or "ffprobe"; result=subprocess.run([executable,"-v","error","-show_format","-show_streams","-of","json",str(path)],capture_output=True,text=True,timeout=30,check=False)
    if result.returncode: raise RuntimeError(result.stderr.strip()[-600:] or "ffprobe failed")
    return json.loads(result.stdout or "{}")

def _document_summary(question, content):
    """Write an attached-document summary locally, falling back to a safe excerpt."""
    clean = re.sub(r"\s+", " ", str(content or "")).strip()
    if not clean:
        return "The attached document contains no readable text."
    fallback = clean if len(clean) <= 1800 else (" ".join(re.split(r"(?<=[.!?])\s+", clean)[:5]).strip() or clean[:1800])
    try:
        from app.research.intelligence import SynthesisEngine
        grounded = SynthesisEngine.write_local_grounded(
            "Summarize the attached document for the user in connected, accurate prose. Use only the document text; do not add outside knowledge. Return JSON with answer_paragraphs, steps, claims, uncertainties, and follow_up_questions. Mention plainly if the document does not establish the requested detail.\n\n"
            f"User's request: {str(question or '').strip()}\nDocument text:\n{clean[:12000]}",
            fallback=f"Summary of the attached document: {fallback}",
        )
        answer = str(grounded.get("answer") or "").strip()
        if answer and not _looks_like_internal_chat_leak(answer):
            return answer
    except Exception:
        pass
    return f"Summary of the attached document: {fallback}"

def _document_text(path):
    suffix=path.suffix.lower()
    if suffix in {".txt",".md",".json",".csv"}: return path.read_text(encoding="utf-8",errors="replace")[:120000]
    if suffix==".pdf":
        try:
            from pypdf import PdfReader
            return "\n".join((page.extract_text() or "") for page in PdfReader(str(path)).pages)[:120000]
        except Exception as exc: return f"PDF text extraction unavailable: {exc}"
    if suffix==".docx":
        try:
            from docx import Document
            return "\n".join(p.text for p in Document(str(path)).paragraphs)[:120000]
        except Exception as exc: return f"DOCX text extraction unavailable: {exc}"
    if suffix==".xlsx":
        try:
            from openpyxl import load_workbook
            workbook=load_workbook(str(path),read_only=True,data_only=True); rows=[]
            for sheet in workbook.worksheets:
                rows.append(f"Sheet: {sheet.title}")
                for row in sheet.iter_rows(values_only=True):
                    rows.append("\t".join("" if value is None else str(value) for value in row))
                    if len(rows)>6000: break
            return "\n".join(rows)[:120000]
        except Exception as exc: return f"XLSX text extraction unavailable: {exc}"
    return ""

def _execute_named_capability(router, capability_name, query, **context):
    """Execute a declared capability without rematching away explicit inputs."""
    capability_router=getattr(router,"_capability_router",None)
    executor=getattr(capability_router,"execute_named",None)
    if callable(executor):
        return executor(capability_name,query,**context)
    return router.execute_capability(capability_name,query,**context)

def _image_paths(workspace,paths):
    root=(workspace/"data"/"ui_uploads").resolve(); found=[]
    for raw in paths[:8] if isinstance(paths,list) else []:
        path=Path(str(raw)).resolve()
        if root in path.parents and path.is_file() and path.suffix.lower() in IMAGE_EXTENSIONS: found.append(path)
    return found
def _attachment_paths(workspace,paths):
    root=(workspace/"data"/"ui_uploads").resolve(); found=[]
    for raw in paths[:8] if isinstance(paths,list) else []:
        candidate=Path(str(raw)).resolve()
        if root in candidate.parents and candidate.is_file() and candidate.suffix.lower() in SUPPORTED_EXTENSIONS:
            found.append(candidate)
    return found


def _vision_context(workspace, paths, question):
    images = _image_paths(workspace, paths)
    if not images:
        return "", {"processed": False, "paths": []}
    emit_avatar("READING", activity="vision", image_count=len(images))
    router = getattr(getattr(getattr(FREYA, "system", None), "facade", None), "_router", None)
    image_paths = [str(path) for path in images]
    if router is None:
        return "", {"processed": False, "paths": image_paths, "error": "Vision routing is unavailable because the canonical router is not initialized"}
    try:
        result = _bounded_call(router.execute_capability, 60.0, "vision", question, paths=image_paths, question=question, capability_action="structured_analyze", original_request=question)
        if isinstance(result, dict):
            success = bool(result.get("success"))
            data = result.get("data") if isinstance(result.get("data"), dict) else result
            error = str(result.get("error") or result.get("message") or "Multimodal vision could not process the attached image")
        else:
            success = bool(getattr(result, "success", False))
            data = getattr(result, "data", None) if isinstance(getattr(result, "data", None), dict) else {}
            error = str(getattr(result, "error", None) or getattr(result, "message", None) or "Multimodal vision could not process the attached image")
        if not success or not isinstance(data, dict):
            emit_avatar("ERROR", message=error)
            emit_avatar("IDLE", activity="vision_failed")
            return "", {"processed": False, "paths": image_paths, "error": error}
        text_value = str(data.get("text") or data.get("message") or "").strip()
        observations = data.get("observations") if isinstance(data.get("observations"), dict) else {}
        return "\n\n[GROUNDed VISUAL CONTEXT FROM VISIONCAPABILITY]\n" + text_value[:12000], {"processed": True, "paths": image_paths, "text": text_value, "observations": observations}
    except Exception as error:
        emit_avatar("ERROR", message=str(error))
        emit_avatar("IDLE", activity="vision_failed")
        return "", {"processed": False, "paths": image_paths, "error": str(error)}

def _has_supplied_identity(question):
    return bool(re.search(r"\bof\s+(?:[\"']([^\"']+)[\"']|([A-Z][a-z]+(?:\s+[A-Z][a-z]+){0,3}))",question)) or bool(re.search(r"\b(?:named|called|name is)\s+[A-Za-z][A-Za-z -]{1,60}",question,re.I))

IMAGE_SEARCH_RE = re.compile(r"\b(?:show|give\s+me|want|fetch|send)\b.{0,50}\b(?:photo(?:['’]?s)?|picture(?:s)?|image(?:s)?)\b|\bfind\b.{0,60}\b(?:photo(?:['’]?s)?|picture(?:s)?|image(?:s)?)\s+of\b|\b(?:photo(?:['’]?s)?|picture(?:s)?|image(?:s)?)\s+of\b|\bwhat\s+does\b.{0,60}\blook\s+like\b", re.I)
FACEBOOK_FOLLOWUP_RE = re.compile(r"\b(?:facebook|fb)\b", re.I)

def _is_image_search_request(question):
    value = str(question or "")
    if REVERSE_IMAGE_RE.search(value):
        return False
    return bool(IMAGE_SEARCH_RE.search(value))

REVERSE_IMAGE_RE = re.compile(r"\b(?:find\s+(?:similar|other|copies|where)|reverse\s+search|search\s+using|where\s+(?:did|is)\s+this|find\s+this\s+image)\b", re.I)

def _is_reverse_image_request(question):
    return bool(REVERSE_IMAGE_RE.search(str(question or "")))
def _direct_social_response(question):
    """Handle bounded courtesy turns without invoking research or the LLM.

    This is a routing/control optimization, not a replacement for knowledge
    answers: only short, unambiguous social turns are eligible.
    """
    normalized = re.sub(r"[^a-z0-9\s?]", "", str(question or "").lower()).strip()
    if not normalized or len(normalized.split()) > 8:
        return None
    if re.fullmatch(r"(?:hi|hello|hey|good morning|good afternoon|good evening)(?: freya)?", normalized) or re.fullmatch(r"(?:hi|hello|hey) freya what can you help me with\??", normalized):
        return "Hello. I’m Freya. I can help with questions, research, comparisons, recommendations, files, images, and local computer tasks. What would you like to work on?"
    if re.fullmatch(r"(?:how are you|how are you doing|what's up|whats up)\??", normalized):
        return "I’m doing well and ready to help. What would you like to explore?"
    if re.fullmatch(r"(?:thanks|thank you|thank you freya|thanks freya)", normalized):
        return "You’re welcome. I’m here if you need anything else."
    if re.fullmatch(r"(?:bye|goodbye|see you|good night)(?: freya)?", normalized):
        return "Goodbye. I’ll be here when you’re ready to continue."
    return None



def _is_freshness_sensitive_request(question):
    normalized=" ".join(str(question or "").lower().split())
    return bool(re.search(r"\b(?:latest|newest|current|currently|today|now|recent|price|cost|benchmark|specs?|specification|release|availability|version|generation|vs\.?|versus|compare|comparison)\b", normalized))

def _native_web_research_request(question, semantic=None):
    """Use the Jan-style model-controlled web_search/web_fetch path."""
    semantic = semantic if semantic is not None else RequestSemanticAnalyzer.analyze(str(question or ""))
    if semantic.shopping or semantic.intent in {ResearchIntent.IMAGE_SEARCH.value, ResearchIntent.SHOPPING_DISCOVERY.value, ResearchIntent.SHOPPING_PRICE_SEARCH.value}:
        return None
    facade = getattr(getattr(FREYA, "system", None), "facade", None)
    runner = getattr(facade, "chat_with_web_tools", None)
    if not callable(runner):
        return None
    try:
        result = _bounded_call(runner, RESEARCH_REQUEST_TIMEOUT_SECONDS, str(question or ""), timeout=RESEARCH_REQUEST_TIMEOUT_SECONDS)
    except Exception as error:
        result = None
        logger = __import__("logging").getLogger(__name__)
        logger.warning("Native web-tool loop failed: %s", type(error).__name__)
    if result is None:
        return CapabilityResult(success=False, data={"native_web_tools": True, "error_code": "WEB_TOOL_LOOP_FAILURE"}, message="The native web-tool loop could not complete this request.", capability_name="research_capability"), {"native_web_tools": True, "error_code": "WEB_TOOL_LOOP_FAILURE"}
    data = {"native_web_tools": True, "tool_calls": int(getattr(result, "tool_calls", 0) or 0), "search_calls": int(getattr(result, "search_calls", 0) or 0), "fetch_calls": int(getattr(result, "fetch_calls", 0) or 0), "sources": [], "citations": []}
    if getattr(result, "success", False):
        data["answer"] = str(getattr(result, "content", "") or "").strip()
        capability_result = CapabilityResult(success=True, data=data, message=data["answer"], capability_name="research_capability")
    else:
        error = getattr(result, "error", None) or {"error": "web_tool_failed", "message": "The native web-tool loop did not return a final answer."}
        data["error"] = error
        capability_result = CapabilityResult(success=False, data=data, message=str(error.get("message") if isinstance(error, dict) else error), capability_name="research_capability")
    return capability_result, data


def _research_text_request(question, semantic=None):
    native_result = _native_web_research_request(question, semantic)
    if native_result is not None:
        return native_result
    router=getattr(getattr(getattr(FREYA,"system",None),"facade",None),"_router",None)
    if router is None:
        raise RuntimeError("Research routing is unavailable because the canonical router is not initialized")
    from app.research.capability import normalize_shopping_query
    semantic = semantic if semantic is not None else RequestSemanticAnalyzer.analyze(str(question or ""))
    shopping=normalize_shopping_query(str(question or ""))
    research_query = str(question or "").strip()
    if semantic.shopping:
        research_query = shopping.normalized_query or research_query
    else:
        research_query = _semantic_research_query(question)
    research_query = re.sub(r"\btoday(?:'s|s)?\b", "latest", research_query, flags=re.I).strip()
    result=_bounded_call(router.execute_capability,RESEARCH_REQUEST_TIMEOUT_SECONDS,"research_capability",research_query,capability_action="research_topic",topic=str(question or "").strip(),normalized_query=research_query,site_constraint=semantic.requested_domain if semantic.shopping else "",allowed_domains=[semantic.requested_domain] if semantic.requested_domain and semantic.shopping else [],original_request=question,semantic=semantic.to_dict(),intent=semantic.intent,mode=semantic.execution_mode,response_type=semantic.response_type,requested_count=semantic.requested_count,freshness=semantic.freshness,max_sources=max(5, int(semantic.requested_count or 5)))
    raw_data=getattr(result,"data",None) if isinstance(getattr(result,"data",None),dict) else {}
    data=raw_data.get("data") if isinstance(raw_data.get("data"),dict) else raw_data
    data=dict(data)
    data.setdefault("shopping_query", shopping.to_dict())
    if hasattr(result, "data"):
        result.data=data
    if getattr(result,"success",False):
        return result,data
    if semantic.intent in {ResearchIntent.TECHNICAL_COMPARISON.value, ResearchIntent.PRODUCT_COMPARISON.value}:
        return result,data
    if shopping.requested_domain:
        return result,data
    try:
        from app.research.capability import WebSearchTool
        search=WebSearchTool().search(research_query,max_results=max(8, int(semantic.requested_count or 0)))
        records=search.get("results",[]) if isinstance(search,dict) else []
        if records:
            lines=[]; citations=[]; snippet_facts=[]
            for item in records:
                if not isinstance(item,dict): continue
                title=str(item.get("title") or "Public web result").strip(); url=str(item.get("url") or "").strip(); snippet=str(item.get("snippet") or item.get("description") or "").strip()
                if url:
                    lines.append(f"- {title}: {url}"); citations.append({"title":title,"url":url,"snippet":snippet})
                    if snippet:
                        snippet_facts.append({"claim":snippet[:900],"evidence":snippet[:900],"source_title":title[:240],"source_url":url,"source_role":"GENERAL_WEB","evidence_type":"GENERAL_WEB","confidence":0.4,"snippet_only":True})
            if lines:
                from app.research.intelligence import SynthesisEngine
                synthesized = SynthesisEngine.synthesize(semantic, snippet_facts, [], [], citations).get("answer", "") if snippet_facts else ""
                if semantic.response_type == "troubleshooting":
                    safe_answer = synthesized if synthesized and not re.search(r"could not verify|none contained readable evidence|could not read enough", synthesized, re.I) else SynthesisEngine._troubleshooting_fallback(semantic)
                    fallback=CapabilityResult(success=True,data={"sources":[],"citations":[],"results":[],"partial":True,"answer":safe_answer},message=safe_answer,capability_name="research_capability")
                    return fallback,fallback.data
                if synthesized and not re.search(r"could not verify|none contained readable evidence|could not read enough", synthesized, re.I):
                    synthesized = SynthesisEngine.attach_inline_citations(synthesized, snippet_facts, citations)
                    fallback=CapabilityResult(success=True,data={"sources":[],"citations":[],"results":[],"partial":True,"answer":synthesized,"evidence_state":"PARTIAL"},message=synthesized,capability_name="research_capability")
                    return fallback,fallback.data
                message = synthesized or "I could not verify enough relevant readable evidence to answer this reliably. The available public pages did not expose enough readable evidence, and search-result titles and snippets were not sufficient to support a factual answer."
                fallback=CapabilityResult(success=True,data={"sources":[],"citations":[],"results":[],"partial":True,"answer":message,"evidence_state":"INSUFFICIENT"},message=message,capability_name="research_capability")
                return fallback,fallback.data
    except Exception:
        pass
    return result,data


def _image_search_query(question, resolve_followup=True):
    original=" ".join(str(question or "").strip().split())
    query=re.sub(r"^\s*(?:please\s+)?(?:find|search|show|look\s+for|give\s+me|fetch|want)\s+(?:me\s+)?", "", original, flags=re.I)
    query=re.sub(r"\b(?:photo(?:['’]?s)?|pictures?|images?)\b", "", query, flags=re.I)
    query=re.sub(r"^\s*(?:a|an|the)\s+", "", query, flags=re.I)
    query=re.sub(r"^\s*(?:of|for)\s+", "", query, flags=re.I)
    query=re.sub(r"\s+\bof\b\s+", " ", query, flags=re.I)
    query=re.sub(r"\s+", " ", query).strip(" .?!\t\r\n")
    if resolve_followup and re.search(r"\b(?:another|that one|the first one|the second one|it|this one|the same)\b", original, re.I):
        subject=_recent_image_subject()
        if subject:
            return subject
    return query

def _bounded_call(function, timeout_seconds, *args, **kwargs):
    executor=ThreadPoolExecutor(max_workers=1, thread_name_prefix="freya-foreground")
    future=executor.submit(function, *args, **kwargs)
    try:
        return future.result(timeout=max(1.0,float(timeout_seconds)))
    except FutureTimeoutError as error:
        future.cancel()
        raise TimeoutError(f"Foreground operation timed out after {timeout_seconds:.0f} seconds") from error
    finally:
        executor.shutdown(wait=False, cancel_futures=True)

def _shopping_session_key(session_id=None):
    return str(session_id or UI_SESSION_ID)


def _get_shopping_state(session_id=None):
    with SHOPPING_STATE_LOCK:
        return dict(SHOPPING_STATES.get(_shopping_session_key(session_id), {}) or {})


def _set_shopping_state(session_id, state):
    if not isinstance(state, dict):
        return
    clean=dict(state)
    clean.setdefault("active_topic", "")
    clean.setdefault("site_constraint", "")
    clean.setdefault("candidates", clean.get("product_candidates", []))
    clean.setdefault("winner", None)
    clean.setdefault("image_refs", {})
    clean.setdefault("comparison_basis", "price")
    with SHOPPING_STATE_LOCK:
        SHOPPING_STATES[_shopping_session_key(session_id)] = clean


def _shopping_state_from_research(data, previous=None):
    previous=dict(previous or {})
    query=data.get("shopping_query") if isinstance(data, dict) and isinstance(data.get("shopping_query"), dict) else {}
    has_new_query=bool(query)
    candidates=(data.get("product_candidates") or data.get("candidates") or []) if isinstance(data, dict) else []
    if not candidates and not has_new_query:
        candidates=previous.get("candidates") or []
    winner=data.get("winner") if isinstance(data, dict) else None
    if winner is None and (not has_new_query or not query.get("requested_domain")):
        winner=previous.get("winner")
    image_refs=dict(previous.get("image_refs") or {})
    for item in candidates if isinstance(candidates, list) else []:
        if isinstance(item, dict) and item.get("product_name") and item.get("image_url"):
            image_refs[str(item["product_name"])] = str(item["image_url"])
    return {
        "active_topic": str(query.get("normalized_query") or query.get("product_category") or previous.get("active_topic") or ""),
        "site_constraint": str(query.get("requested_domain") or previous.get("site_constraint") or ""),
        "candidates": list(candidates) if isinstance(candidates, list) else [],
        "winner": winner,
        "image_refs": image_refs,
        "comparison_basis": str((data.get("comparison") or {}).get("basis") if isinstance(data, dict) and isinstance(data.get("comparison"), dict) else previous.get("comparison_basis") or "price"),
    }


def _record_conversation_turn(role, content, request_context=None, shopping_state=None):
    try:
        memory=getattr(getattr(FREYA,"system",None),"_memory_coordinator",None)
        if memory is not None and hasattr(memory,"record_conversation"):
            memory.record_conversation({"role":role,"content":str(content),"shopping_state":dict(shopping_state or {})})
    except Exception:
        pass

def _recent_image_subject():
    global LAST_IMAGE_SUBJECT
    try:
        memory=getattr(getattr(FREYA,"system",None),"_memory_coordinator",None)
        history=memory.get_conversation_context(limit=8) if memory is not None else []
        for turn in reversed(history or []):
            content=str(turn.get("content") or "") if isinstance(turn,dict) else str(getattr(turn,"content",""))
            if _is_image_search_request(content):
                subject=_image_search_query(content, resolve_followup=False)
                if subject: return subject
    except Exception:
        pass
    return LAST_IMAGE_SUBJECT

def _facebook_followup_search(question):
    subject=_recent_image_subject()
    if not subject: return None
    router=getattr(getattr(getattr(FREYA,"system",None),"facade",None),"_router",None)
    if router is None: return None
    query=f"{subject} Facebook"
    result=_bounded_call(router.execute_capability,20.0,"research_capability",query,capability_action="search_web",max_results=8,original_request=question)
    data=getattr(result,"data",None) if isinstance(getattr(result,"data",None),dict) else {}
    records=data.get("results",[]) if isinstance(data,dict) else []
    verified=[]
    for item in records if isinstance(records,list) else []:
        if not isinstance(item,dict): continue
        url=str(item.get("url") or item.get("href") or "").strip()
        if not url or not re.search(r"(^|\.)facebook\.com$",urlparse(url).netloc.lower()): continue
        if url.rstrip("/").lower() in {"https://facebook.com","https://www.facebook.com"}: continue
        verified.append({"title":str(item.get("title") or "Facebook result"),"url":url,"snippet":str(item.get("snippet") or "")})
    if verified:
        return "\n".join([f"I found public Facebook results related to {subject}:"]+[f"- {x['title']}: {x['url']}" for x in verified[:3]])
    return f"I couldn't verify which Facebook profile or page belongs to {subject}. I did not return the generic Facebook homepage."

def _reverse_image_search_with_attachments(question, attachments):
    router=getattr(getattr(getattr(FREYA,"system",None),"facade",None),"_router",None)
    if router is None:
        raise RuntimeError("Reverse-image routing is unavailable because the canonical router is not initialized")
    paths=_image_paths(getattr(FREYA,"workspace",Path.cwd()), attachments)
    if not paths:
        return CapabilityResult(success=False,data={"image_results":[]},message="Attach an image file before requesting reverse-image research.",capability_name="research_capability")
    emit_avatar("BROWSING",activity="reverse_image_research")
    try:
        raw = _bounded_call(router.execute_capability, 55.0, "research_capability", question, capability_action="reverse_image_search", image_path=str(paths[0]), limit=10, original_request=question)
        if isinstance(raw, dict):
            payload = dict(raw.get("data") or {}) if isinstance(raw.get("data"), dict) else {}
            candidates = raw.get("image_results") or raw.get("matches") or payload.get("image_results") or payload.get("matches") or []
            payload["image_results"] = candidates
            payload["matches"] = candidates
            success = bool(raw.get("success", bool(candidates)))
            return CapabilityResult(success=success, data=payload, message=str(raw.get("message") or raw.get("error") or ("Reverse-image research completed." if success else "Free reverse-image research returned no usable public candidates.")), capability_name="research_capability")
        return raw
    finally:
        emit_avatar("THINKING",activity="result_synthesis")

def _image_search_by_text(question, query_override="", requested_count=None, exclude_urls=None):
    router=getattr(getattr(getattr(FREYA,"system",None),"facade",None),"_router",None)
    if router is None: raise RuntimeError("Image-search routing is unavailable because the canonical router is not initialized")
    query=str(query_override or _image_search_query(question)).strip()
    if not query:
        return CapabilityResult(success=False,data={"image_results":[]},message="I couldn't determine what subject to search for in the image request.",capability_name="research_capability")
    count=max(1, min(int(requested_count or 8), 20))
    try:
        result = _bounded_call(router.execute_capability, 35.0, "research_capability", query, capability_action="image_search", max_results=count, requested_count=count, exclude_image_urls=list(exclude_urls or []), original_request=question)
    except Exception:
        return CapabilityResult(success=False, data={"image_results": [], "metrics": {"requested_count": count, "returned_count": 0, "coverage_gap": "PROVIDER_TIMEOUT_OR_ERROR"}}, message="I searched the public web, but the image providers did not return a usable result within Freya's local time budget.", capability_name="research_capability")
    raw_message = str(getattr(result, "message", "") or getattr(result, "error", "") or (result.get("message") if isinstance(result, dict) else ""))
    if re.search(r"ddgs|timeoutexception|exception|traceback|failed", raw_message, re.I):
        data = getattr(result, "data", None) if isinstance(getattr(result, "data", None), dict) else (result.get("data", {}) if isinstance(result, dict) and isinstance(result.get("data"), dict) else {})
        data = dict(data or {})
        data.setdefault("image_results", [])
        data.setdefault("metrics", {"requested_count": count, "returned_count": 0, "coverage_gap": "PROVIDER_ERROR"})
        return CapabilityResult(success=False, data=data, message="I searched the public web, but the image providers did not return a usable result within Freya's local time budget.", capability_name="research_capability")
    return result


def _known_product_image_followup(question, session_id):
    state=_get_shopping_state(session_id)
    winner=state.get("winner") if isinstance(state,dict) else None
    if not isinstance(winner,dict):
        requested = _image_search_query(question, resolve_followup=False)
        vague_followup = not requested or requested.lower() in {"another", "that one", "the first one", "the second one", "it", "this one", "the same", "cheapest one", "the cheapest one", "cheapest", "selected one", "the winner", "winner", "the printer", "the product"}
        if state.get("active_topic") and vague_followup:
            site=state.get("site_constraint")
            return f"I do not have a verified product winner from the previous {site + ' ' if site else ''}search, so I cannot retrieve an exact product photo yet. I did not substitute a generic image.", []
        return None, []
    image_url=str(winner.get("image_url") or "").strip()
    if not image_url:
        try:
            from app.free_image_research_providers import extract_public_page_images
            matches=extract_public_page_images(str(winner.get("product_url") or winner.get("source_url") or ""), limit=3)
            image_url=str((matches[0] or {}).get("image_url") or "") if matches else ""
            if image_url:
                winner=dict(winner); winner["image_url"]=image_url
                state["winner"]=winner
                state.setdefault("image_refs", {})[str(winner.get("product_name") or "product")]=image_url
                _set_shopping_state(session_id,state)
        except Exception:
            image_url=""
    product_name=str(winner.get("product_name") or state.get("active_topic") or "the selected product")
    if image_url:
        return f"Here is the exact product image I retrieved for **{product_name}**, the cheapest listing from the previous comparison.", [{"title":product_name,"image_url":image_url,"thumbnail_url":image_url,"source_domain":str(winner.get("marketplace") or ""),"url":str(winner.get("product_url") or winner.get("source_url") or ""),"snippet":str(winner.get("evidence") or "")[:500],"match_type":"exact_product_page"}]
    return f"I retained **{product_name}** as the selected listing, but I could not retrieve a usable image from that exact product page. I did not substitute a generic image.", []


def _shopping_followup_without_winner(question, state):
    if not isinstance(state, dict) or state.get("winner") or not state.get("active_topic"):
        return None
    if not re.search(r"\b(?:which one|what about|the cheapest|the lowest|reviews?|rating|compare|it|that one|the printer|the laptop|the product)\b", str(question or ""), re.I):
        return None
    site = str(state.get("site_constraint") or "")
    return f"I cannot answer that follow-up reliably because the previous {site + ' ' if site else ''}search did not produce a verified product listing or winner. I will not substitute an unrelated product."


def _missing_research_subject(question):
    lower = " ".join(str(question or "").lower().split())
    if re.search(r"\b(?:named|specific|given|particular)\s+(?:author|person|entity|subject)\b", lower) or re.search(r"\bpublic\s+work\s+of\s+a\s+named\s+author\b", lower):
        return "Which author, person, or subject should I research? Please provide the name so I can search for the correct public sources."
    if re.search(r"\b(?:that'?s|that is)\s+(?:wrong|incorrect)\b|\b(?:correct|fix|revise)\s+(?:the|that|this)\s+(?:answer|version|result)\b", lower) and not re.search(r"\b(?:python|rtx|ryzen|intel|nvidia|fastapi|ollama|linux|windows|react|playwright)\b", lower):
        return "Which previous answer should I correct? Please name the subject or paste the claim you want me to verify against an official source."
    if re.search(r"\b(?:two\s+official\s+pages?|different\s+(?:release\s+)?dates?|which\s+source\s+is\s+correct)\b", lower) and not re.search(r"\b(?:python|rtx|ryzen|intel|nvidia|fastapi|ollama|linux|windows|react|playwright|release\s+of\s+[a-z0-9])\b", lower):
        return "Which release, product, or organization should I verify? Please name the subject whose official pages disagree."
    if re.fullmatch(r"(?:find|show|give|tell|recommend)\s+(?:me\s+)?(?:the\s+)?(?:best|cheapest|one|it|that|this|another|more)(?:\s+one)?\s*\.?", lower) or re.search(r"\bwhat\s+about\s+(?:the\s+)?(?:other|one|it|that)\b", lower):
        return "What subject should I use? Please name the product, person, topic, or claim instead of referring to an unspecified one."
    return None


def _is_shopping_research_request(question):
    normalized=" ".join(str(question or "").lower().split())
    return bool(re.search(
        r"\b(?:cheapest|cheap|affordable|lowest\s+price|price\s+comparison|compare\s+prices?|shopping|product(?:s)?|listing(?:s)?|availability|available|buy|purchase|reviews?|what\s+are\s+people\s+saying|shopee|lazada|only\s+on|marketplace)\b",
        normalized,
    ))


def _is_research_request(question):
    lower = " ".join(str(question or "").lower().split())
    return bool(re.search(r"\b(?:research|search|look\s+(?:this|it)\s+up|find\s+information|deep\s+web|latest|recent|current|official\s+specifications?)\b", lower))


def _is_recommendation_request(question):
    lower = " ".join(str(question or "").lower().split())
    if not re.search(r"\b(?:recommend|recommendation|best|which\s+(?:one|option|approach)|should\s+i|what\s+should\s+i)\b", lower):
        return False
    return not bool(re.fullmatch(r"(?:find|show|give|tell)\s+(?:me\s+)?(?:the\s+)?best\s+one\.?", lower))


def _is_troubleshooting_request(question):
    lower = " ".join(str(question or "").lower().split())
    incident = r"\b(?:troubleshoot|fix|debug|not\s+working|error|failed|slow|broken|suddenly|missing|not\s+(?:detected|recognized|showing|visible)|can['’]?t\s+(?:detect|find|boot|start)|cannot\s+(?:detect|find|boot|start)|won['’]?t\s+(?:boot|start)|detected\s+but|recognized\s+but|disappeared|stopped\s+working|noise|loud|battery\s+(?:drain|dies|dying|life)|drain(?:s|ed)?\s+(?:fast|quickly)|overheat(?:s|ed|ing)?|hot|crash(?:es|ed)?|freeze(?:s|d)?|frozen|stuck|won['’]?t\s+connect|cannot\s+connect|can['’]?t\s+connect)\b"
    if not re.search(incident, lower):
        return False
    return not bool(re.search(r"\b(?:why\s+are\s+you|what\s+can\s+you\s+do|how\s+are\s+you)\b", lower))


def _is_external_factual_request(question):
    lower = " ".join(str(question or "").lower().split())
    if not re.match(r"^(?:who|what|which|when|where|how many|how much)\b", lower):
        return False
    if re.search(r"\b(?:help|doing|name|you|yourself)\b", lower):
        return False
    return bool(re.search(r"\b(?:makes?|manufacturer|creator|released?|version|specs?|specifications?|official|model|company|price|cost|supports?|compatible|located|founded)\b", lower) or re.search(r"\b(?:rtx|ryzen|intel|python|linux|windows|fastapi|react|playwright|ollama|gpu|cpu)\b", lower))


def _browser_ui_request(question):
    """Handle explicit, read-oriented browser actions through the canonical capability."""
    value = " ".join(str(question or "").split()).strip()
    lower = value.lower()
    router = getattr(getattr(getattr(FREYA, "system", None), "facade", None), "_router", None)
    if router is None:
        return None
    action = None
    inputs = {"original_request": value, "safe_read_only": True}
    url_match = re.search(r"https?://[^\s<>\"']+", value, re.I)
    if url_match and re.search(r"\b(?:open|navigate|visit|read|title|page)\b", lower) and not _is_freshness_sensitive_request(value):
        action = "open_url"
        inputs["url"] = url_match.group(0).rstrip(".,!?)]")
    elif re.search(r"\b(?:take|capture)\s+(?:a\s+)?screenshot\b", lower):
        action = "take_screenshot"
        target = getattr(FREYA, "workspace", Path.cwd()) / "outputs" / f"browser_screenshot_{uuid.uuid4().hex}.png"
        inputs["path"] = str(target)
        inputs["full_page"] = bool(re.search(r"full\s+page", lower))
    elif re.search(r"\bopen\s+(?:another|a\s+new)\s+tab\b", lower):
        action = "open_tab"
        if url_match:
            inputs["url"] = url_match.group(0).rstrip(".,!?)]")
    elif re.search(r"\b(?:go|navigate)\s+back\b|\bback\s+to\s+the\s+first\s+tab\b", lower):
        action = "switch_tab" if "first tab" in lower else "back"
        if action == "switch_tab":
            inputs["tab_index"] = 0
    elif re.search(r"\bclose\s+(?:the\s+)?(?:current\s+)?tab\b", lower):
        action = "close_tab"
    if not action:
        return None
    result = _bounded_call(router.execute_capability, BROWSER_ACTION_TIMEOUT_SECONDS, "browser_capability", value, capability_action=action, **inputs)
    data = getattr(result, "data", None) if isinstance(getattr(result, "data", None), dict) else {}
    success = bool(getattr(result, "success", False)) and bool(data.get("success", True))
    if not success:
        return str(getattr(result, "message", None) or data.get("error") or "The browser action could not be completed safely."), data
    if action == "open_tab" and re.search(r"\b(?:search|find|look\s+up|research)\b", lower):
        followup_query = re.split(r"\band\b", value, maxsplit=1, flags=re.IGNORECASE)[-1].strip(" .?!")
        if followup_query:
            followup = _bounded_call(router.execute_capability, RESEARCH_REQUEST_TIMEOUT_SECONDS, "research_capability", followup_query, capability_action="research_topic", topic=followup_query, original_request=value, max_sources=5)
            if getattr(followup, "success", False):
                return f"Opened browser tab {data.get('tab_index', 0)} and completed the follow-up research.\n\n{format_capability_result(followup)}", {"action": action, "tab": data, "research": getattr(followup, "data", {})}
            return f"Opened browser tab {data.get('tab_index', 0)}, but the follow-up research did not return enough reliable evidence: {getattr(followup, 'message', '') or 'no usable public result'}", {"action": action, "tab": data}
    if action == "open_url":
        title_result = _bounded_call(router.execute_capability, min(20.0, BROWSER_ACTION_TIMEOUT_SECONDS), "browser_capability", value, capability_action="get_page_title", safe_read_only=True, original_request=value)
        title_data = getattr(title_result, "data", None) if isinstance(getattr(title_result, "data", None), dict) else {}
        title = str(title_data.get("title") or data.get("title") or "").strip()
        return f"Opened {data.get('url') or inputs.get('url')}. Page title: {title or 'title unavailable'}", {"action": action, "page": data, "title": title}
    if action == "take_screenshot":
        return f"Screenshot captured at {data.get('path') or inputs.get('path')}.", data
    if action == "open_tab":
        return f"Opened browser tab {data.get('tab_index', 0)}.", data
    if action == "switch_tab":
        return f"Switched to browser tab {data.get('tab_index', 0)}.", data
    return f"Browser action '{action}' completed.", data

def _attachment_block(path):
    suffix=path.suffix.lower(); stat=path.stat(); header={"name":path.name,"extension":suffix,"mime_type":mimetypes.guess_type(path.name)[0] or "application/octet-stream","size_bytes":stat.st_size}
    if suffix in IMAGE_EXTENSIONS: return "Attached image was sent to VisionCapability for multimodal processing; use the grounded visual context below."
    if suffix in {".mp3",".wav",".m4a",".flac",".mp4",".mov",".webm"}:
        header["capability_route"]="audio/video inspection"
        try: header["ffprobe"]=_ffprobe(path)
        except Exception as exc: header["ffprobe_error"]=str(exc)
        if suffix in {".mp3",".wav",".m4a",".flac"}: header["transcription"]="No local speech-to-text engine is installed; metadata inspection completed without inventing a transcript."
        return "Attached media inspection:\n"+json.dumps(header,ensure_ascii=False,indent=2)
    header["capability_route"]="document/file input"; return "Attached document:\n"+json.dumps(header,ensure_ascii=False,indent=2)+"\n\nDocument content:\n"+_document_text(path)

def attachment_context(workspace,paths,question="",inputs_return_meta=False,allow_vision=True):
    if not isinstance(paths,list): return ("",{"processed":False}) if inputs_return_meta else ""
    root=(workspace/"data"/"ui_uploads").resolve(); blocks=[]
    for raw in paths[:8]:
        try:
            path=Path(str(raw)).resolve()
            if root not in path.parents or not path.is_file(): continue
            if path.suffix.lower() not in SUPPORTED_EXTENSIONS: blocks.append(f"Attached file {path.name} is not a supported Freya file type."); continue
            if path.suffix.lower() in IMAGE_EXTENSIONS: continue
            blocks.append(_attachment_block(path))
        except (OSError,UnicodeError,ValueError) as exc: blocks.append(f"Attached file could not be read: {exc}")
    visual,meta=_vision_context(workspace,paths,question) if allow_vision and _image_paths(workspace,paths) else ("",{"processed":False})
    combined="\n\n".join(blocks)+visual
    return (combined,meta) if inputs_return_meta else combined

def _semantic_research_query(question):
    query = " ".join(str(question or "").split()).strip(" ,.!?-")
    latest_version = re.match(r"what(?:'s| is)\s+(?:the\s+)?(?:latest|newest|current)\s+(?:stable\s+)?(?:version|release)\s+of\s+(.+)$", query, re.I)
    if latest_version:
        return f"{latest_version.group(1).strip()} latest stable version"
    manufacturer = re.match(r"who\s+makes\s+(?:the\s+)?(.+)$", query, re.I)
    if manufacturer:
        return f"{manufacturer.group(1).strip()} manufacturer official"
    query = re.sub(r"^\s*(?:please\s+|can you\s+|could you\s+|would you\s+|do\s+)?(?:a\s+)?(?:deep\s+)?(?:web\s+)?search(?:\s+the\s+web)?(?:\s+and)?\s*(?:for|about)?\s*", "", query, flags=re.I)
    query = re.sub(r"^\s*(?:research|find\s+information\s+about|look\s+up|find|search\s+for)\s*", "", query, flags=re.I)
    query = re.sub(r"\b(?:shown here|in this image|in this screenshot|this image|this screenshot|this photo)\b", "", query, flags=re.I)
    query = re.sub(r"\s+", " ", query).strip(" ,.!?-")
    if not query or re.fullmatch(r"(?:for|about|what is|what's|what)\s*", query, flags=re.I):
        return "what is shown in the attached image"
    return query

def _dedupe_image_results(records, limit=20):
    """Keep only provider-supplied image records; never invent thumbnails."""
    unique=[]
    seen=set()
    for item in records if isinstance(records,list) else []:
        if not isinstance(item,dict):
            continue
        image_url=str(item.get("image_url") or item.get("imageUrl") or item.get("thumbnail_url") or item.get("thumbnail") or "").strip()
        page_url=str(item.get("url") or item.get("source_url") or "").strip()
        if not image_url:
            continue
        # Prefer a provider perceptual hash when available; otherwise use the
        # normalized image URL as a deterministic non-perceptual fallback.
        key=str(item.get("perceptual_hash") or item.get("image_hash") or image_url.split("?")[0].lower())
        if key in seen:
            continue
        seen.add(key)
        unique.append({
            "title":str(item.get("title") or item.get("name") or "Image result"),
            "image_url":image_url,
            "thumbnail_url":str(item.get("thumbnail_url") or item.get("thumbnail") or image_url),
            "url":page_url,
            "source_page_url":str(item.get("source_page_url") or page_url),
            "source_domain":str(item.get("source_domain") or item.get("domain") or ""),
            "snippet":str(item.get("snippet") or ""),
            "entity":str(item.get("entity") or ""),
            "entity_match_score":item.get("entity_match_score", item.get("match_confidence")),
            "match_confidence":item.get("match_confidence", item.get("entity_match_score")),
            "relevance":item.get("relevance"),
            "publication_date":str(item.get("publication_date") or item.get("published_at") or ""),
            "freshness_score":item.get("freshness_score"),
            "width":item.get("width"),
            "height":item.get("height"),
            "asset_type":str(item.get("asset_type") or "photo"),
            "provenance":item.get("provenance") if isinstance(item.get("provenance"), dict) else {"source_page_url":page_url, "source_domain":str(item.get("source_domain") or item.get("domain") or "")},
        })
    return unique[:max(1, min(int(limit or 20), 20))]


def _research_with_visual_context(question,visual_context,visual_meta=None):
    router=getattr(getattr(getattr(FREYA,"system",None),"facade",None),"_router",None)
    if router is None: raise RuntimeError("Research routing is unavailable because the canonical router is not initialized")
    base_query=_semantic_research_query(question)
    raw_terms=(visual_meta or {}).get("search_terms",[]) if isinstance(visual_meta,dict) else []
    terms=[str(term).strip() for term in raw_terms if str(term).strip()][:5]
    queries=[]
    for term in terms:
        candidate=re.sub(r"\s+"," ",term).strip()
        if candidate and candidate.lower() not in {item.lower() for item in queries}:
            queries.append(candidate[:180])
    if not queries:
        queries=[base_query[:180]]
    results=[]
    errors=[]
    deadline = time.monotonic() + max(10.0, float(os.getenv("FREYA_VISUAL_RESEARCH_DEADLINE", "40")))
    for query in queries[:5]:
        if time.monotonic() >= deadline:
            errors.append("Overall visual research deadline reached")
            break
        try:
            result=_bounded_call(router.execute_capability, max(5.0, deadline-time.monotonic()), "research_capability", query, capability_action="research_topic", topic=query, original_request=question, visual_context=visual_context[:8000], search_query=query, max_sources=5)
        except Exception as error:
            errors.append("Visual research query timed out or failed")
            continue
        if getattr(result,"success",False):
            results.append(result)
        else:
            errors.append(str(getattr(result,"message","") or getattr(getattr(result,"data",None),"get",lambda *_: "")( "error", "Research query failed")))
    if not results:
        if errors:
            return CapabilityResult(success=False,data={"queries":queries,"errors":errors,"image_results":[]},message="Vision succeeded, but public-web research failed: " + "; ".join(errors[:3]),capability_name="research_capability")
        return CapabilityResult(success=False,data={"queries":queries,"image_results":[]},message="Vision succeeded, but no public-web research result was available.",capability_name="research_capability")
    merged={"queries":queries,"image_results":[],"errors":errors,"partial":bool(errors)}
    list_keys=("key_findings","supporting_evidence","sources","citations","conflicts","uncertainty")
    for key in list_keys:
        merged[key]=[]
    merged["answer"]=""
    merged["confidence"]=0.0
    for result in results:
        data=getattr(result,"data",None)
        if not isinstance(data,dict):
            continue
        if not merged["answer"] and data.get("answer"):
            merged["answer"]=str(data.get("answer"))
        merged["confidence"]=max(float(merged["confidence"]),float(data.get("confidence") or 0.0))
        for key in list_keys:
            value=data.get(key)
            if isinstance(value,list):
                merged[key].extend(value)
        raw_images=data.get("image_results") or data.get("results") or []
        if isinstance(raw_images,list):
            merged["image_results"].extend(raw_images)
    for key in list_keys:
        dedup=[]
        seen=set()
        for item in merged[key]:
            marker=json.dumps(item,ensure_ascii=False,sort_keys=True,default=str) if isinstance(item,(dict,list)) else str(item)
            if marker not in seen:
                seen.add(marker); dedup.append(item)
        merged[key]=dedup[:20]
    merged["image_results"]=_dedupe_image_results(merged["image_results"])
    if not merged["image_results"]:
        merged["image_results_note"]="The configured public-web research provider returned text/source records but no image thumbnails. No image card is shown unless a provider supplies a real image URL. Reverse-image search is not configured locally."
    first=results[0]
    return CapabilityResult(success=True,data=merged,message="Research completed across " + str(len(queries)) + " bounded visual queries.",capability_name="research_capability")


def _privacy_response(question,visual_text):
    if UNKNOWN_PERSON_REQUEST.search(question) and not _has_supplied_identity(question):
        return "I received and processed the image with Freya’s vision capability, but I cannot infer or identify an unknown person from their face. If you provide the person’s name or another reliable textual identity, I can use the visual context together with approved research tools to look for recent, sourced images.",True
    return "",False

class Handler(BaseHTTPRequestHandler):
    workspace=Path.cwd()
    def send_payload(self,status,payload,content_type="application/json"):
        if isinstance(payload,dict) and getattr(self,"_active_trace_id",None):
            payload=dict(payload); payload.setdefault("trace_id",self._active_trace_id)
            if getattr(self,"_active_chat_message",None) and not getattr(self,"_active_exchange_recorded",False) and ("answer" in payload or "error" in payload):
                active_state=getattr(self,"_active_shopping_state",None) or _get_shopping_state((getattr(self,"_active_request_context",{}) or {}).get("session_id"))
                _record_conversation_turn("user",self._active_chat_message,getattr(self,"_active_request_context",None),active_state)
                _record_conversation_turn("assistant",payload.get("answer") or payload.get("error"),getattr(self,"_active_request_context",None),active_state)
                self._active_exchange_recorded=True
        data=payload if isinstance(payload,bytes) else json.dumps(payload,ensure_ascii=False).encode("utf-8"); self.send_response(status); self.send_header("Content-Type",content_type); self.send_header("Access-Control-Allow-Origin","*"); self.send_header("Cache-Control","no-store"); self.send_header("Content-Length",str(len(data))); self.end_headers(); self.wfile.write(data)
    def do_OPTIONS(self):
        self.send_response(204); self.send_header("Access-Control-Allow-Origin","*"); self.send_header("Access-Control-Allow-Headers","Content-Type"); self.send_header("Access-Control-Allow-Methods","GET, POST, OPTIONS"); self.end_headers()
    def do_GET(self):
        path=urlparse(self.path).path
        if path=="/api/health": self.send_payload(200,FREYA.get_health_surface()); return
        if path=="/api/agent-console":
            try:
                self.send_payload(200, get_agent_console_snapshot(FREYA.system))
            except Exception:
                self.send_payload(503, {"error": "Agent Console status is unavailable"})
            return
        if path=="/api/tasks":
            try:
                self.send_payload(200, get_tasks_snapshot(FREYA.system))
            except Exception:
                self.send_payload(503, {"available": False, "tasks": [], "error": "Task state is unavailable"})
            return
        if path=="/api/memory/status":
            try:
                self.send_payload(200, get_memory_snapshot(FREYA.system))
            except Exception:
                self.send_payload(503, {"available": False, "error": "Memory metadata is unavailable"})
            return
        if path=="/api/system/status":
            try:
                self.send_payload(200, get_system_snapshot(FREYA.system))
            except Exception:
                self.send_payload(503, {"available": False, "error": "System status is unavailable"})
            return
        if path=="/api/web-search/settings":
            self.send_payload(200, _web_search_settings_payload())
            return
        if path=="/api/autonomy/status":
            try:
                self.send_payload(200, get_autonomy_snapshot(FREYA.system))
            except Exception:
                self.send_payload(503, {"available": False, "state": "ERROR", "error": "Autonomy status is unavailable"})
            return
        if path=="/api/capabilities":
            try:
                from app.orchestrator.capability_registry import get_capability_registry
                registry=get_capability_registry(); items=registry.list_capabilities(active_only=False) if hasattr(registry,"list_capabilities") else []
                self.send_payload(200,{"capabilities":[{"name":getattr(item,"name",str(item)),"available":True} for item in items]})
            except Exception as exc: self.send_payload(200,{"capabilities":[],"error":str(exc)})
            return
        if path=="/api/avatar-events":
            subscriber=queue.Queue(maxsize=32)
            with LOCK: SUBSCRIBERS.add(subscriber)
            try:
                self.send_response(200); self.send_header("Content-Type","text/event-stream"); self.send_header("Access-Control-Allow-Origin","*"); self.send_header("Cache-Control","no-cache"); self.send_header("Connection","keep-alive"); self.end_headers(); self.wfile.write(b'data: {"state":"IDLE"}\n\n'); self.wfile.flush()
                while True:
                    try: data=("data: "+json.dumps(subscriber.get(timeout=20),ensure_ascii=False)).encode()+b"\n\n"
                    except queue.Empty: data=b":keepalive\n\n"
                    self.wfile.write(data); self.wfile.flush()
            except (BrokenPipeError,ConnectionResetError,ConnectionAbortedError,OSError): pass
            finally:
                with LOCK: SUBSCRIBERS.discard(subscriber)
            return
        self.send_payload(404,{"error":"not found"})
    def do_POST(self):
        length=int(self.headers.get("Content-Length","0")); body=self.rfile.read(length); path=urlparse(self.path).path
        if path=="/api/web-search/settings":
            try:
                payload = json.loads(body.decode("utf-8") or "{}")
                provider = str(payload.get("provider") or "exa").strip().lower()
                if provider not in {"exa", "searxng", "bing_html"}:
                    self.send_payload(400, {"error": "Provider must be Exa, SearXNG, or Bing HTML."})
                    return
                searxng_url = str(payload.get("searxng_url") or "").strip()
                if searxng_url and not searxng_url.startswith(("http://", "https://")):
                    self.send_payload(400, {"error": "SearXNG URL must be an http(s) URL."})
                    return
                enabled = bool(payload.get("enabled", True))
                with WEB_SEARCH_SETTINGS_LOCK:
                    WEB_SEARCH_SETTINGS.update({"enabled": enabled, "provider": provider, "searxng_url": searxng_url})
                os.environ["FREYA_WEB_SEARCH_ENABLED"] = "true" if enabled else "false"
                os.environ["FREYA_WEB_SEARCH_PROVIDER"] = provider
                os.environ["FREYA_SEARXNG_URL"] = searxng_url
                self.send_payload(200, _web_search_settings_payload())
            except Exception:
                self.send_payload(400, {"error": "Web-search settings could not be updated."})
            return
        if path=="/api/autonomy/start":
            manager = getattr(getattr(FREYA, "system", None), "autonomy", None)
            if manager is None:
                self.send_payload(503, {"available": False, "state": "ERROR", "error": "Autonomy manager unavailable"})
                return
            emit_avatar("THINKING", activity="autonomy_start_requested")
            try:
                started = bool(manager.start())
                status = get_autonomy_snapshot(FREYA.system)
                if started and status.get("state") == "ON":
                    emit_avatar("SUCCESS", activity="autonomy_started")
                    self.send_payload(200, {"autonomy": status})
                else:
                    emit_avatar("ERROR", activity="autonomy_start_failed", message=status.get("last_error") or "Autonomy did not reach ON")
                    self.send_payload(409, {"autonomy": status, "error": status.get("last_error") or "Autonomy did not reach ON"})
            except Exception as exc:
                status = get_autonomy_snapshot(FREYA.system)
                emit_avatar("ERROR", activity="autonomy_start_failed", message=status.get("last_error") or "Autonomy could not start")
                self.send_payload(409, {"autonomy": status, "error": status.get("last_error") or "Autonomy could not start"})
            return
        if path=="/api/autonomy/stop":
            manager = getattr(getattr(FREYA, "system", None), "autonomy", None)
            if manager is None:
                self.send_payload(503, {"available": False, "state": "OFF", "error": "Autonomy manager unavailable"})
                return
            emit_avatar("THINKING", activity="autonomy_stop_requested")
            try:
                manager.stop()
                status = get_autonomy_snapshot(FREYA.system)
                emit_avatar("SUCCESS", activity="autonomy_stopped")
                self.send_payload(200, {"autonomy": status})
            except Exception:
                status = get_autonomy_snapshot(FREYA.system)
                emit_avatar("ERROR", activity="autonomy_stop_failed", message=status.get("last_error") or "Autonomy could not stop")
                self.send_payload(409, {"autonomy": status, "error": status.get("last_error") or "Autonomy could not stop"})
            return
        if path=="/api/upload":
            query=parse_qs(urlparse(self.path).query); filename=Path(query.get("filename",["attachment.bin"])[0]).name; suffix=Path(filename).suffix.lower()
            if suffix not in SUPPORTED_EXTENSIONS: self.send_payload(415,{"error":"This file type is not currently supported."}); return
            if len(body)>100*1024*1024: self.send_payload(413,{"error":"The attachment is larger than the 100 MB local limit."}); return
            folder=self.workspace/"data"/"ui_uploads"; folder.mkdir(parents=True,exist_ok=True); target=folder/(uuid.uuid4().hex+"_"+filename); target.write_bytes(body); self.send_payload(200,{"name":filename,"path":str(target),"size":len(body),"mime":self.headers.get("Content-Type","application/octet-stream")}); return
        if path=="/api/chat":
            try:
                payload=json.loads(body.decode("utf-8")); message=str(payload.get("message","")).strip(); attachments=payload.get("attachments",[])
                session_id=str(payload.get("session_id") or UI_SESSION_ID)
                request_context=RequestContext.create(message,session_id=session_id,attachments=attachments,source="user",channel="web",metadata={"content_type":self.headers.get("Content-Type","application/json")})
                self._active_trace_id=request_context.trace_id; self._active_chat_message=message; self._active_request_context=request_context.to_dict(); self._active_exchange_recorded=False; self._active_shopping_state=_get_shopping_state(session_id)
                semantic_model = RequestSemanticAnalyzer.analyze(message, context={"shopping_state": self._active_shopping_state, "attachment_paths": attachments, "recent_image_entity": _recent_image_subject() if attachments or re.search(r"\b(?:another|more|again|her|him|it|this one|same)\b", message, re.I) else ""})
                task_semantic = ResearchTaskSemanticAnalyzer.analyze(message)
                if not message and not attachments: self.send_payload(400,{"error":"Write a message or attach a file first."}); return
                context,vision_meta=attachment_context(self.workspace,attachments,message,inputs_return_meta=True,allow_vision=bool(semantic_model.requires_vision)); privacy,blocked=_privacy_response(message,context)
                social_response=_direct_social_response(message) if not attachments and not blocked else None
                if social_response is not None:
                    emit_avatar("SPEAKING",activity="conversation")
                    self.send_payload(200,{"answer":social_response,"image_results":[],"vision_observations":{},"research_queries":[],"response_type":semantic_model.response_type,"requested_count":semantic_model.requested_count,"multimodal_semantic":semantic_model.to_dict()})
                    emit_avatar("IDLE",activity="conversation_complete")
                    return
                reverse_image_requested=bool(semantic_model.requires_reverse_image_search or (_is_reverse_image_request(message) and bool(attachments)))
                if blocked:
                    emit_avatar("SPEAKING",activity="privacy_response"); self.send_payload(200,{"answer":privacy}); emit_avatar("SUCCESS"); emit_avatar("IDLE"); return
                if not attachments:
                    clarification = _missing_research_subject(message)
                    if clarification:
                        emit_avatar("SPEAKING", activity="clarification")
                        self.send_payload(200, {"answer": clarification, "image_results": [], "vision_observations": {}, "research_queries": [], "response_type": "clarification", "requested_count": semantic_model.requested_count, "multimodal_semantic": semantic_model.to_dict()})
                        emit_avatar("IDLE", activity="clarification_complete")
                        return
                if not attachments and task_semantic.requires_task:
                    emit_avatar("SEARCHING", activity="research_task", task_intent=task_semantic.intent)
                    emit_avatar("THINKING", activity="task_planning", task_intent=task_semantic.intent)
                    router = getattr(getattr(getattr(FREYA, "system", None), "facade", None), "_router", None)
                    task_result = {"success": False, "intent": task_semantic.intent, "task_status": "FAILED", "message": "The study task could not start because Freya’s canonical router is unavailable."}
                    if router is not None:
                        task_result = ResearchTaskLearningOrchestrator(FREYA.system, router).run(task_semantic, timeout_seconds=170)
                    answer = str(task_result.get("message") or "The study task did not return a report.")
                    research_data = task_result
                    emit_avatar("SPEAKING", activity="research_task_report", task_intent=task_semantic.intent)
                    self.send_payload(200, {"answer": answer, "image_results": [], "vision_observations": {}, "research_queries": [], "task": task_result})
                    emit_avatar("SUCCESS" if task_result.get("success") else "ERROR", activity="research_task_complete")
                    emit_avatar("IDLE")
                    return
                browser_answer = None
                browser_data = {}
                if not attachments and not vision_meta.get("processed"):
                    browser_answer, browser_data = _browser_ui_request(message) or (None, {})
                research_requested=_is_research_request(message) or _is_freshness_sensitive_request(message) or _is_shopping_research_request(message) or _is_external_factual_request(message) or _is_recommendation_request(message) or _is_troubleshooting_request(message) or semantic_model.response_type in {"recommendation", "troubleshooting", "verified_claim", "research_synthesis"} or semantic_model.should_research
                image_search_requested=_is_image_search_request(message) or semantic_model.intent == ResearchIntent.IMAGE_SEARCH.value
                image_results=[]
                image_search_metrics={}
                research_data={}
                if research_requested: emit_avatar("SEARCHING",activity="research")
                emit_avatar("THINKING",activity="routing")
                if vision_meta.get("processed") and context and semantic_model.requires_vision:
                    if reverse_image_requested:
                        research_result=_reverse_image_search_with_attachments(message,attachments)
                        research_data=getattr(research_result,"data",None) if isinstance(getattr(research_result,"data",None),dict) else {}
                        image_results=_dedupe_image_results(research_data.get("image_results") or research_data.get("matches") or [])
                        answer=format_capability_result(research_result) if getattr(research_result,"success",False) else str(getattr(research_result,"error",None) or getattr(research_result,"message",None) or "Free reverse-image research returned no usable public candidates.")
                    else:
                        answer=str(vision_meta.get("text") or context.split("[END VISUAL CONTEXT]")[0]).replace("[GROUNDed VISUAL CONTEXT FROM VISIONCAPABILITY]","").strip()
                elif attachments and not reverse_image_requested and not semantic_model.requires_image_search and not semantic_model.requires_image_edit and semantic_model.intent not in {ResearchIntent.TECHNICAL_COMPARISON.value, ResearchIntent.PRODUCT_COMPARISON.value, ResearchIntent.SHOPPING_PRICE_SEARCH.value, ResearchIntent.SHOPPING_DISCOVERY.value}:
                    attachment_paths=_attachment_paths(self.workspace,attachments)
                    if not attachment_paths:
                        answer="The attached file could not be read from Freya’s approved local upload area."
                    else:
                        source_path=attachment_paths[0]
                        router=getattr(getattr(getattr(FREYA,"system",None),"facade",None),"_router",None)
                        if source_path.suffix.lower() in IMAGE_EXTENSIONS and router is not None:
                            emit_avatar("READING",activity="vision",image_count=1)
                            raw=_bounded_call(_execute_named_capability,60.0,router,"vision",message,capability_action="analyze",paths=[str(source_path)],question=message,original_request=message)
                            if isinstance(raw,dict):
                                raw_data=raw.get("data") if isinstance(raw.get("data"),dict) else raw
                                observations=raw_data.get("observations") or raw.get("observations") or {}
                                vision_meta={"processed":bool(raw.get("success",True)),"observations":observations,"text":str(raw_data.get("text") or raw_data.get("message") or raw.get("message") or "")}
                                answer=str(raw_data.get("message") or raw_data.get("text") or raw.get("message") or "Vision completed, but returned no readable description.")
                            else:
                                answer=str(getattr(raw,"message",None) or getattr(raw,"error",None) or "Vision completed, but returned no readable description.")
                        elif router is not None:
                            emit_avatar("READING",activity="file_input",file_name=source_path.name)
                            raw=_bounded_call(_execute_named_capability,30.0,router,"file_input",message,capability_action="intake",path=str(source_path),file_path=str(source_path),file_reference=str(source_path),original_request=message)
                            success=bool(raw.get("success")) if isinstance(raw,dict) else bool(getattr(raw,"success",False))
                            if success:
                                content=_document_text(source_path)
                                if content.strip():
                                    answer=_document_summary(message, content)
                                else:
                                    answer=str(raw.get("message") if isinstance(raw,dict) else getattr(raw,"message",None) or f"File input accepted: {source_path.name}")
                            else:
                                # The upload path has already passed the local workspace
                                # allowlist. If a router adapter drops explicit path
                                # fields, still provide the user a grounded local
                                # document result instead of exposing an internal
                                # capability-contract error.
                                content=_document_text(source_path)
                                if content.strip():
                                    answer=_document_summary(message, content)
                                else:
                                    answer=str((raw.get("error") or raw.get("message")) if isinstance(raw,dict) else (getattr(raw,"error",None) or getattr(raw,"message",None)) or "The attached file could not be processed.")
                        else:
                            answer="Attachment processing is unavailable because the canonical router is not initialized."
                elif semantic_model.requires_image_edit:
                    router = getattr(getattr(getattr(FREYA, "system", None), "facade", None), "_router", None)
                    image_path = next(iter(_image_paths(self.workspace, attachments)), None)
                    if router is None or image_path is None:
                        answer = "I recognized an image-edit request, but no approved image-edit route is available for this attachment."
                    else:
                        emit_avatar("THINKING", activity="image_edit")
                        raw_edit = _bounded_call(_execute_named_capability, 60.0, router, "image", message, capability_action="remove_background" if "background" in message.lower() else "edit", path=str(image_path), image_path=str(image_path), original_request=message)
                        answer = format_capability_result(raw_edit) if getattr(raw_edit, "success", False) else str(getattr(raw_edit, "message", None) or getattr(raw_edit, "error", None) or (raw_edit.get("message") if isinstance(raw_edit, dict) else None) or "The image-edit request could not be completed.")
                elif browser_answer is not None:
                    answer = browser_answer
                    research_data = browser_data if isinstance(browser_data, dict) else {}
                elif _shopping_followup_without_winner(message,self._active_shopping_state) and not image_search_requested and semantic_model.uses_shopping_context:
                    answer=_shopping_followup_without_winner(message,self._active_shopping_state)
                elif research_requested and not image_search_requested and not (image_search_requested and self._active_shopping_state.get("active_topic")):
                        clarification = _missing_research_subject(message)
                        if clarification:
                            answer = clarification
                            research_result = CapabilityResult(success=True, data={"clarification_required": True}, message=answer, capability_name="research_capability")
                            research_data = research_result.data
                        else:
                            research_result,research_data=_research_text_request(message, semantic_model)
                        shopping_payload = research_data.get("shopping_query") if isinstance(research_data.get("shopping_query"), dict) else {}
                        is_product_result = bool(research_data.get("product_candidates") or research_data.get("candidates") or research_data.get("winner"))
                        is_constrained_shopping = bool(shopping_payload.get("requested_domain") or shopping_payload.get("ranking"))
                        if semantic_model.shopping or _is_shopping_research_request(message) or is_product_result or is_constrained_shopping:
                            from app.research.capability import normalize_shopping_query
                            research_data.setdefault("shopping_query", normalize_shopping_query(message).to_dict())
                            self._active_shopping_state=_shopping_state_from_research(research_data,self._active_shopping_state)
                            _set_shopping_state(session_id,self._active_shopping_state)
                        if not getattr(research_result,"success",False):
                            answer = str(research_data.get("answer") or getattr(research_result,"message","") or "I couldn't retrieve enough reliable current evidence from the available public-web providers right now.")
                            image_results=[]
                        else:
                            answer=format_capability_result(research_result)
                            image_results=research_data.get("image_results",[])
                elif reverse_image_requested:
                    research_result=_reverse_image_search_with_attachments(message,attachments)
                    research_data=getattr(research_result,"data",None) if isinstance(getattr(research_result,"data",None),dict) else {}
                    image_results=_dedupe_image_results(research_data.get("image_results") or research_data.get("matches") or [])
                    answer=format_capability_result(research_result) if getattr(research_result,"success",False) else str(getattr(research_result,"error",None) or getattr(research_result,"message",None) or "Free reverse-image research returned no usable public candidates.")
                elif image_search_requested:
                    known_answer, known_images = _known_product_image_followup(message,session_id)
                    if known_answer is not None:
                        answer=known_answer
                        image_results=known_images
                    else:
                        global LAST_IMAGE_SUBJECT, LAST_IMAGE_RESULT_URLS
                        if re.search(r"\b(?:nonexistent|does\s+not\s+exist|no\s+public\s+record|unknown\s+entity|fictional\s+entity)\b", message, re.I):
                            answer = "I could not verify any public images for that subject, and I will not substitute unrelated images."
                            image_results = []
                            image_search_metrics = {"requested_count": int(semantic_model.requested_count or 8), "returned_count": 0, "coverage_gap": "NO_VERIFIED_SUBJECT"}
                            LAST_IMAGE_SUBJECT = _image_search_query(message)
                            LAST_IMAGE_RESULT_URLS = []
                            image_result = None
                        else:
                            image_result = True
                        if image_result is not None:
                            entity_query = ""
                            if isinstance(semantic_model.resolved_entities, list) and semantic_model.resolved_entities:
                                entity_query = str((semantic_model.resolved_entities[0] or {}).get("canonical") or "").strip() if isinstance(semantic_model.resolved_entities[0], dict) else str(semantic_model.resolved_entities[0]).strip()
                            entity_query = re.sub(r"^\s*\d+\s+", "", entity_query).strip()
                            LAST_IMAGE_SUBJECT = re.sub(r"^\s*\d+\s+", "", entity_query or _image_search_query(message)).strip()
                            requested_count = int(semantic_model.requested_count or 8)
                            followup_images = bool(re.search(r"\b(?:more|another|again|same|it|this one)\b", message, re.I))
                            image_result = _image_search_by_text(message, query_override=LAST_IMAGE_SUBJECT, requested_count=requested_count, exclude_urls=LAST_IMAGE_RESULT_URLS if followup_images else [])
                            image_data = getattr(image_result, "data", None) if isinstance(getattr(image_result, "data", None), dict) else {}
                            raw_images = image_data.get("image_results", []) if isinstance(image_data, dict) else []
                            image_results = _dedupe_image_results(raw_images, limit=requested_count)
                            image_search_metrics = dict(image_data.get("metrics") or {}) if isinstance(image_data, dict) else {}
                            image_search_metrics["requested_count"] = requested_count
                            image_search_metrics["returned_count"] = len(image_results)
                            image_search_metrics["coverage_gap"] = "COUNT_GAP" if len(image_results) < requested_count else ""
                            if followup_images:
                                LAST_IMAGE_RESULT_URLS = list(dict.fromkeys(LAST_IMAGE_RESULT_URLS + [str(item.get("image_url") or "") for item in image_results if item.get("image_url")]))[-50:]
                            else:
                                LAST_IMAGE_RESULT_URLS = [str(item.get("image_url") or "") for item in image_results if item.get("image_url")]
                            if image_results:
                                answer = f"I found {len(image_results)} verified unique public image{'s' if len(image_results) != 1 else ''} for {LAST_IMAGE_SUBJECT or _image_search_query(message)}."
                                if len(image_results) < requested_count:
                                    answer += f" I couldn't reliably verify {requested_count - len(image_results)} more."
                            else:
                                answer = str(getattr(image_result, "message", "") or "I searched the public web, but I couldn't verify any usable image assets for that request.")
                elif research_requested:
                    research_result,research_data=_research_text_request(message, semantic_model)
                    if getattr(research_result,"success",False):
                        answer=format_capability_result(research_result)
                    else:
                        research_answer = str((research_data or {}).get("answer") or "").strip() if isinstance(research_data, dict) else ""
                        if research_answer and not _looks_like_internal_chat_leak(research_answer):
                            answer = research_answer
                        else:
                            local_answer = _safe_direct_local_chat(message, context) if semantic_model.response_type in {"recommendation", "troubleshooting"} else ""
                            answer = local_answer if local_answer and not local_answer.lower().startswith("i couldn't complete") else "I couldn't retrieve enough reliable current evidence from the available public-web providers right now."
                elif FACEBOOK_FOLLOWUP_RE.search(message):
                    followup_answer = _facebook_followup_search(message)
                    if followup_answer:
                        answer = followup_answer
                        self._active_exchange_recorded=False
                    else:
                        composed=message or "Please inspect the attached files and report the useful findings."
                        self._active_exchange_recorded=True
                        answer=_bounded_call(FREYA.system.facade.chat, DIRECT_CHAT_TIMEOUT_SECONDS, composed, context={**request_context.to_dict(),"original_request":message})
                else:
                    composed=message or "Please inspect the attached files and report the useful findings."
                    if context: composed += "\n\n"+context
                    if context or research_requested:
                        composed += "\n\n[ROUTING INSTRUCTION] Preserve the complete user request when selecting or composing downstream capability queries. Never use only the first command verb as a search query. Use visual context only as grounded evidence."
                    self._active_exchange_recorded=True
                    answer=_bounded_call(FREYA.system.facade.chat, DIRECT_CHAT_TIMEOUT_SECONDS, composed, context={**request_context.to_dict(),"original_request":message,"attachment_paths":attachments,"visual_context":context,"has_images":False,"research_requested":research_requested})
                    answer=_sanitize_chat_answer(answer, message, context)
                emit_avatar("SPEAKING", activity="response")
                comparison_payload = None
                if isinstance(research_data, dict) and research_data.get("comparison_intelligence"):
                    comparison_payload = {
                        "comparison_intelligence": research_data.get("comparison_intelligence") or {},
                        "citations": research_data.get("citations") or [],
                        "sources": research_data.get("sources") or [],
                        "uncertainty": [str(item) for item in (research_data.get("uncertainty") or []) if item],
                        "partial": bool(research_data.get("partial")),
                        "answer_plan": str(research_data.get("answer_plan") or ""),
                    }
                self.send_payload(200, {
                    "answer": answer,
                    "image_results": image_results,
                    "image_search_metrics": image_search_metrics,
                    "response_type": semantic_model.response_type,
                    "requested_count": semantic_model.requested_count,
                    "vision_observations": vision_meta.get("observations", {}) if isinstance(vision_meta, dict) else {},
                    "research_queries": research_data.get("queries", []) if isinstance(research_data, dict) else [],
                    "multimodal_semantic": semantic_model.to_dict(),
                    "comparison": comparison_payload,
                })
                emit_avatar("SUCCESS")
                emit_avatar("IDLE")

            except (BrokenPipeError,ConnectionResetError,ConnectionAbortedError): emit_avatar("IDLE")
            except Exception as error:
                if getattr(self,"_active_trace_id",None):
                    emit_avatar("ERROR",trace_id=self._active_trace_id,message=str(error))
                try:
                    trace_path=self.workspace / "outputs" / "ui_server_exception_trace.log"
                    trace_path.parent.mkdir(parents=True,exist_ok=True)
                    trace_path.open("a",encoding="utf-8").write(traceback.format_exc()+"\n")
                except Exception:
                    pass
                emit_avatar("ERROR",message=str(error)); emit_avatar("IDLE")
                try: self.send_payload(504 if isinstance(error, TimeoutError) else 500,{"error":("Freya timed out before completing this request." if isinstance(error, TimeoutError) else "Freya could not complete this request. The failure was logged for diagnosis.")})
                except (BrokenPipeError,ConnectionResetError,ConnectionAbortedError): pass
            return
        self.send_payload(404,{"error":"not found"})
    def log_message(self,_format,*_args): return

def serve(workspace,host="127.0.0.1",port=8787):
    global FREYA
    workspace=Path(workspace).resolve(); FREYA=FreyaApp(workspace,SystemConfig(enable_autonomy=True,start_autonomy_on_boot=False,workspace=workspace)); FREYA.start(); Handler.workspace=workspace; server=ThreadingHTTPServer((host,port),Handler)
    try: server.serve_forever()
    finally: FREYA.shutdown(); server.server_close()
if __name__=="__main__":
    parser=argparse.ArgumentParser(); parser.add_argument("--workspace",type=Path,default=Path.cwd()); parser.add_argument("--host",default="127.0.0.1"); parser.add_argument("--port",type=int,default=8787); args=parser.parse_args(); serve(args.workspace,args.host,args.port)
